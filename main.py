from fake_useragent import UserAgent
from openpyxl.styles import PatternFill
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
from selenium.common.exceptions import TimeoutException, NoSuchElementException



ua = UserAgent(os='windows', browsers=['edge', 'chrome'], min_percentage=1.3)
random_user_agent = ua.random

# Keep the browser open after the program finishes
options = Options()
options.add_experimental_option("detach", True)
options.add_argument(f"user-agent={random_user_agent}")

driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
driver.get("https://www.ampto.com/")
driver.maximize_window()
file_path = "AMPTO LIST PRICE 2024 IN EXCEL (include net map and close out).xlsx"



def readData(file, sheetName, rowNo, columnNo):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.cell(rowNo, columnNo).value


def rowCount(file, sheetName):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetName]
    return sheet.max_row



workbook = openpyxl.Workbook()
sheet = workbook.active

sheet.append(["Models", "Images", "Names", "Descriptions", "Links"])


rows = rowCount(file_path, "Sheet")

for data in range(2, rows + 1):
    try:
        code = readData(file_path, "Sheet1", data, 1)
        input_element = WebDriverWait(driver, 10).until(EC.visibility_of_element_located((By.CSS_SELECTOR, "input[placeholder='Search...']")))

        input_element.send_keys(code)
        driver.find_element(By.XPATH, "//button[@aria-label='Search']").click()

        wait = WebDriverWait(driver, 10, poll_frequency=2)

        container = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, ".product-list.product-list--collection")))

        first_link = container.find_element(By.XPATH, ".//a")

        link_url = first_link.get_attribute("href")

        driver.execute_script("arguments[0].scrollIntoView({behavior: 'auto', block: 'center'});", first_link)

        first_link.click()

        # find the image
        try:
            image = driver.find_element(By.CSS_SELECTOR, ".product-gallery__image").get_attribute("src")
        except NoSuchElementException:
            image = "not available"

        # find the name
        try:
            name = driver.find_element(By.CSS_SELECTOR, ".product-meta__title.heading.h1").text
        except NoSuchElementException:
            name = "not available"

        # find the description
        try:
            description = driver.find_element(By.CSS_SELECTOR, "div[class='rte text--pull'] p").text
        except NoSuchElementException:
            description = "not available"

        # Print the URL, name, description, and image
        print("Link URL:", link_url)
        print("name:", name)
        print("description:", description)
        print("image:", image)
        print("")

        # Append the data to the Excel sheet
        sheet.append([code, image, name, description, link_url])

        workbook.save("output.xlsx")



    except TimeoutException:
        print("Search results not available for code:", code)
        sheet.append([code, "not available", "not available", "not available", "not available"])
        workbook.save("output.xlsx")
        continue
    except NoSuchElementException:
        print("Some elements not found for code:", code)
        sheet.append([code, "not available", "not available", "not available", "not available"])
        workbook.save("output.xlsx")

        continue

workbook.close()
driver.quit()